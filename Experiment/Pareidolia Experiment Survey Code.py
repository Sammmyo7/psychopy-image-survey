#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
This experiment was created using PsychoPy3 Experiment Builder (v2023.2.3),
    on Tue May  7 23:34:01 2024
If you publish work using this script the most relevant publication is:

    Peirce J, Gray JR, Simpson S, MacAskill M, Höchenberger R, Sogo H, Kastman E, Lindeløv JK. (2019) 
        PsychoPy2: Experiments in behavior made easy Behav Res 51: 195. 
        https://doi.org/10.3758/s13428-018-01193-y

"""

# --- Import packages ---
from psychopy import locale_setup
from psychopy import prefs
from psychopy import plugins
plugins.activatePlugins()
prefs.hardware['audioLib'] = 'ptb'
prefs.hardware['audioLatencyMode'] = '0'
from psychopy import sound, gui, visual, core, data, event, logging, clock, colors, layout
from psychopy.tools import environmenttools
from psychopy.constants import (NOT_STARTED, STARTED, PLAYING, PAUSED,
                                STOPPED, FINISHED, PRESSED, RELEASED, FOREVER, priority)

import numpy as np  # whole numpy lib is available, prepend 'np.'
from numpy import (sin, cos, tan, log, log10, pi, average,
                   sqrt, std, deg2rad, rad2deg, linspace, asarray)
from numpy.random import random, randint, normal, shuffle, choice as randchoice
import os  # handy system and path functions
import sys  # to get file system encoding

from psychopy.hardware import keyboard

# --- Setup global variables (available in all functions) ---
# Ensure that relative paths start from the same directory as this script
_thisDir = os.path.dirname(os.path.abspath(__file__))
# Store info about the experiment session
psychopyVersion = '2023.2.3'
expName = 'Pareidolia Experiment Survey'  # from the Builder filename that created this script
expInfo = {
    'participant': f"{randint(0, 999999):06.0f}",
    'session': '001',
    'date': data.getDateStr(),  # add a simple timestamp
    'expName': expName,
    'psychopyVersion': psychopyVersion,
}


def showExpInfoDlg(expInfo):
    """
    Show participant info dialog.
    Parameters
    ==========
    expInfo : dict
        Information about this experiment, created by the `setupExpInfo` function.
    
    Returns
    ==========
    dict
        Information about this experiment.
    """
    # temporarily remove keys which the dialog doesn't need to show
    poppedKeys = {
        'date': expInfo.pop('date', data.getDateStr()),
        'expName': expInfo.pop('expName', expName),
        'psychopyVersion': expInfo.pop('psychopyVersion', psychopyVersion),
    }
    # show participant info dialog
    dlg = gui.DlgFromDict(dictionary=expInfo, sortKeys=False, title=expName)
    if dlg.OK == False:
        core.quit()  # user pressed cancel
    # restore hidden keys
    expInfo.update(poppedKeys)
    # return expInfo
    return expInfo


def setupData(expInfo, dataDir=None):
    """
    Make an ExperimentHandler to handle trials and saving.
    
    Parameters
    ==========
    expInfo : dict
        Information about this experiment, created by the `setupExpInfo` function.
    dataDir : Path, str or None
        Folder to save the data to, leave as None to create a folder in the current directory.    
    Returns
    ==========
    psychopy.data.ExperimentHandler
        Handler object for this experiment, contains the data to save and information about 
        where to save it to.
    """
    
    # data file name stem = absolute path + name; later add .psyexp, .csv, .log, etc
    if dataDir is None:
        dataDir = _thisDir
    filename = u'data/%s_%s_%s' % (expInfo['participant'], expName, expInfo['date'])
    # make sure filename is relative to dataDir
    if os.path.isabs(filename):
        dataDir = os.path.commonprefix([dataDir, filename])
        filename = os.path.relpath(filename, dataDir)
    
    # an ExperimentHandler isn't essential but helps with data saving
    thisExp = data.ExperimentHandler(
        name=expName, version='',
        extraInfo=expInfo, runtimeInfo=None,
        originPath='/Users/morlando24/Documents/Experiment/Pareidolia Experiment Survey.py',
        savePickle=True, saveWideText=True,
        dataFileName=dataDir + os.sep + filename, sortColumns='time'
    )
    thisExp.setPriority('thisRow.t', priority.CRITICAL)
    thisExp.setPriority('expName', priority.LOW)
    # return experiment handler
    return thisExp


def setupLogging(filename):
    """
    Setup a log file and tell it what level to log at.
    
    Parameters
    ==========
    filename : str or pathlib.Path
        Filename to save log file and data files as, doesn't need an extension.
    
    Returns
    ==========
    psychopy.logging.LogFile
        Text stream to receive inputs from the logging system.
    """
    # this outputs to the screen, not a file
    logging.console.setLevel(logging.EXP)


def setupWindow(expInfo=None, win=None):
    """
    Setup the Window
    
    Parameters
    ==========
    expInfo : dict
        Information about this experiment, created by the `setupExpInfo` function.
    win : psychopy.visual.Window
        Window to setup - leave as None to create a new window.
    
    Returns
    ==========
    psychopy.visual.Window
        Window in which to run this experiment.
    """
    if win is None:
        # if not given a window to setup, make one
        win = visual.Window(
            size=[1440, 900], fullscr=False, screen=0,
            winType='pyglet', allowStencil=False,
            monitor='testMonitor', color=[0,0,0], colorSpace='rgb',
            backgroundImage='', backgroundFit='none',
            blendMode='avg', useFBO=True,
            units='height'
        )
        if expInfo is not None:
            # store frame rate of monitor if we can measure it
            expInfo['frameRate'] = win.getActualFrameRate()
    else:
        # if we have a window, just set the attributes which are safe to set
        win.color = [0,0,0]
        win.colorSpace = 'rgb'
        win.backgroundImage = ''
        win.backgroundFit = 'none'
        win.units = 'height'
    win.mouseVisible = True
    win.hideMessage()
    return win


def setupInputs(expInfo, thisExp, win):
    """
    Setup whatever inputs are available (mouse, keyboard, eyetracker, etc.)
    
    Parameters
    ==========
    expInfo : dict
        Information about this experiment, created by the `setupExpInfo` function.
    thisExp : psychopy.data.ExperimentHandler
        Handler object for this experiment, contains the data to save and information about 
        where to save it to.
    win : psychopy.visual.Window
        Window in which to run this experiment.
    Returns
    ==========
    dict
        Dictionary of input devices by name.
    """
    # --- Setup input devices ---
    inputs = {}
    ioConfig = {}
    ioSession = ioServer = eyetracker = None
    
    # create a default keyboard (e.g. to check for escape)
    defaultKeyboard = keyboard.Keyboard(backend='ptb')
    # return inputs dict
    return {
        'ioServer': ioServer,
        'defaultKeyboard': defaultKeyboard,
        'eyetracker': eyetracker,
    }

def pauseExperiment(thisExp, inputs=None, win=None, timers=[], playbackComponents=[]):
    """
    Pause this experiment, preventing the flow from advancing to the next routine until resumed.
    
    Parameters
    ==========
    thisExp : psychopy.data.ExperimentHandler
        Handler object for this experiment, contains the data to save and information about 
        where to save it to.
    inputs : dict
        Dictionary of input devices by name.
    win : psychopy.visual.Window
        Window for this experiment.
    timers : list, tuple
        List of timers to reset once pausing is finished.
    playbackComponents : list, tuple
        List of any components with a `pause` method which need to be paused.
    """
    # if we are not paused, do nothing
    if thisExp.status != PAUSED:
        return
    
    # pause any playback components
    for comp in playbackComponents:
        comp.pause()
    # prevent components from auto-drawing
    win.stashAutoDraw()
    # run a while loop while we wait to unpause
    while thisExp.status == PAUSED:
        # make sure we have a keyboard
        if inputs is None:
            inputs = {
                'defaultKeyboard': keyboard.Keyboard(backend='PsychToolbox')
            }
        # check for quit (typically the Esc key)
        if inputs['defaultKeyboard'].getKeys(keyList=['escape']):
            endExperiment(thisExp, win=win, inputs=inputs)
        # flip the screen
        win.flip()
    # if stop was requested while paused, quit
    if thisExp.status == FINISHED:
        endExperiment(thisExp, inputs=inputs, win=win)
    # resume any playback components
    for comp in playbackComponents:
        comp.play()
    # restore auto-drawn components
    win.retrieveAutoDraw()
    # reset any timers
    for timer in timers:
        timer.reset()


def run(expInfo, thisExp, win, inputs, globalClock=None, thisSession=None):
    """
    Run the experiment flow.
    
    Parameters
    ==========
    expInfo : dict
        Information about this experiment, created by the `setupExpInfo` function.
    thisExp : psychopy.data.ExperimentHandler
        Handler object for this experiment, contains the data to save and information about 
        where to save it to.
    psychopy.visual.Window
        Window in which to run this experiment.
    inputs : dict
        Dictionary of input devices by name.
    globalClock : psychopy.core.clock.Clock or None
        Clock to get global time from - supply None to make a new one.
    thisSession : psychopy.session.Session or None
        Handle of the Session object this experiment is being run from, if any.
    """
    # mark experiment as started
    thisExp.status = STARTED
    # make sure variables created by exec are available globally
    exec = environmenttools.setExecEnvironment(globals())
    # get device handles from dict of input devices
    ioServer = inputs['ioServer']
    defaultKeyboard = inputs['defaultKeyboard']
    eyetracker = inputs['eyetracker']
    # make sure we're running in the directory for this experiment
    os.chdir(_thisDir)
    # get filename from ExperimentHandler for convenience
    filename = thisExp.dataFileName
    frameTolerance = 0.001  # how close to onset before 'same' frame
    endExpNow = False  # flag for 'escape' or other condition => quit the exp
    # get frame duration from frame rate in expInfo
    if 'frameRate' in expInfo and expInfo['frameRate'] is not None:
        frameDur = 1.0 / round(expInfo['frameRate'])
    else:
        frameDur = 1.0 / 60.0  # could not measure, so guess
    
    # Start Code - component code to be run after the window creation
    
    # --- Initialize components for Routine "welcome_screen" ---
    # Run 'Begin Experiment' code from codefix
    # Begin Experiment tab
    from psychopy import visual, core, event
    
    welcometext = visual.TextStim(win=win, name='welcometext',
        text="Welcome to this survey!\nCarefully READ the instructions and respond to the survey prompts accordingly. It's highly recommended that this be completed INDEPENDENTLY in one sitting. ",
        font='Arial',
        pos=(0, 0), height=0.05, wrapWidth=None, ori=0.0, 
        color='white', colorSpace='rgb', opacity=None, 
        languageStyle='LTR',
        depth=-1.0);
    forwardkey = keyboard.Keyboard()
    advance = visual.TextStim(win=win, name='advance',
        text="Press the 'right arrow' on your keyboard to continue. [There is no going back after advancing]",
        font='Arial',
        pos=(0, -0.3), height=0.05, wrapWidth=None, ori=0.0, 
        color='white', colorSpace='rgb', opacity=None, 
        languageStyle='LTR',
        depth=-3.0);
    
    # --- Initialize components for Routine "warning_prep" ---
    warningtime = visual.TextStim(win=win, name='warningtime',
        text='The 10 SECOND countdown will begin on the next screen. \n\nYou will be shown an array of images and you must CHOOSE ONE IMAGE that stands out the MOST to you as QUICKLY as possible within the time limit using mouse input. ',
        font='Arial',
        pos=(0, 0), height=0.05, wrapWidth=None, ori=0.0, 
        color='white', colorSpace='rgb', opacity=None, 
        languageStyle='LTR',
        depth=0.0);
    forwardkey2 = keyboard.Keyboard()
    advance2 = visual.TextStim(win=win, name='advance2',
        text="Press the 'right arrow' on your keyboard to continue. ",
        font='Arial',
        pos=(0, -0.3), height=0.05, wrapWidth=None, ori=0.0, 
        color='white', colorSpace='rgb', opacity=None, 
        languageStyle='LTR',
        depth=-2.0);
    
    # --- Initialize components for Routine "Experiment1_countdown" ---
    OnionGrid = visual.ImageStim(
        win=win,
        name='OnionGrid', 
        image=None, mask=None, anchor='center',
        ori=0.0, pos=(0, 0), size=(0.125, 0.125),
        color=[1,1,1], colorSpace='rgb', opacity=None,
        flipHoriz=False, flipVert=False,
        texRes=128.0, interpolate=True, depth=0.0)
    
    # --- Initialize components for Routine "second_prep" ---
    secondtestwarning = visual.TextStim(win=win, name='secondtestwarning',
        text='In the next part of this survey, an array of images will be displayed for 35 seconds. \n\nPlease observe these images CAREFULLY for the allotted time while waiting for the next instruction.',
        font='Arial',
        pos=(0, 0), height=0.05, wrapWidth=None, ori=0.0, 
        color='white', colorSpace='rgb', opacity=None, 
        languageStyle='LTR',
        depth=0.0);
    forwardkey3 = keyboard.Keyboard()
    advance3 = visual.TextStim(win=win, name='advance3',
        text="Press the 'right arrow' on your keyboard to continue. ",
        font='Arial',
        pos=(0, -0.3), height=0.05, wrapWidth=None, ori=0.0, 
        color='white', colorSpace='rgb', opacity=None, 
        languageStyle='LTR',
        depth=-2.0);
    
    # --- Initialize components for Routine "Experiement2_countdown" ---
    
    # --- Initialize components for Routine "Experiement2_instruction" ---
    selectioninstruction = visual.TextStim(win=win, name='selectioninstruction',
        text="Now, an array of 48 images will be shown. \n\nSelect the 24 IMAGES that YOU REMEMBER seeing in the array previously shown. Answer to the best of your memory. There's no countdown. The experiment won't end until 24 images have been selected. ",
        font='Arial',
        pos=(0,0), height=0.05, wrapWidth=None, ori=0.0, 
        color='white', colorSpace='rgb', opacity=None, 
        languageStyle='LTR',
        depth=0.0);
    forwardkey4 = keyboard.Keyboard()
    advance4 = visual.TextStim(win=win, name='advance4',
        text="Press the 'right arrow' on your keyboard to continue. ",
        font='Arial',
        pos=(0, -0.3), height=0.05, wrapWidth=None, ori=0.0, 
        color='white', colorSpace='rgb', opacity=None, 
        languageStyle='LTR',
        depth=-2.0);
    
    # --- Initialize components for Routine "Experiment2_imageselection" ---
    TreeGrid = visual.ImageStim(
        win=win,
        name='TreeGrid', 
        image=None, mask=None, anchor='center',
        ori=0.0, pos=(0, 0), size=(0.125, 0.125),
        color=[1,1,1], colorSpace='rgb', opacity=None,
        flipHoriz=False, flipVert=False,
        texRes=128.0, interpolate=True, depth=0.0)
    
    # --- Initialize components for Routine "End" ---
    endtext = visual.TextStim(win=win, name='endtext',
        text='Thank you for participating in this survey on Pareidolia. Did you notice the faces?',
        font='Arial',
        pos=(0, 0), height=0.05, wrapWidth=None, ori=0.0, 
        color='white', colorSpace='rgb', opacity=None, 
        languageStyle='LTR',
        depth=0.0);
    endkey = keyboard.Keyboard()
    endadvance = visual.TextStim(win=win, name='endadvance',
        text="Press the 'right arrow' on your keyboard to end and close the survey :D",
        font='Arial',
        pos=(0, -0.25), height=0.05, wrapWidth=None, ori=0.0, 
        color='white', colorSpace='rgb', opacity=None, 
        languageStyle='LTR',
        depth=-2.0);
    
    # create some handy timers
    if globalClock is None:
        globalClock = core.Clock()  # to track the time since experiment started
    if ioServer is not None:
        ioServer.syncClock(globalClock)
    logging.setDefaultClock(globalClock)
    routineTimer = core.Clock()  # to track time remaining of each (possibly non-slip) routine
    win.flip()  # flip window to reset last flip timer
    # store the exact time the global clock started
    expInfo['expStart'] = data.getDateStr(format='%Y-%m-%d %Hh%M.%S.%f %z', fractionalSecondDigits=6)
    
    # --- Prepare to start Routine "welcome_screen" ---
    continueRoutine = True
    # update component parameters for each repeat
    thisExp.addData('welcome_screen.started', globalClock.getTime())
    # Run 'Begin Routine' code from codefix
    welcometext = visual.TextStim(win=win, name='welcometext',
        text=welcometext.text,
        font='Arial',
        pos=(0, 0), height=0.05,
        wrapWidth=None, ori=0, 
        color='white', colorSpace='rgb', opacity=1,
        depth=0.0)
    
    forwardkey.keys = []
    forwardkey.rt = []
    _forwardkey_allKeys = []
    # keep track of which components have finished
    welcome_screenComponents = [welcometext, forwardkey, advance]
    for thisComponent in welcome_screenComponents:
        thisComponent.tStart = None
        thisComponent.tStop = None
        thisComponent.tStartRefresh = None
        thisComponent.tStopRefresh = None
        if hasattr(thisComponent, 'status'):
            thisComponent.status = NOT_STARTED
    # reset timers
    t = 0
    _timeToFirstFrame = win.getFutureFlipTime(clock="now")
    frameN = -1
    
    # --- Run Routine "welcome_screen" ---
    routineForceEnded = not continueRoutine
    while continueRoutine:
        # get current time
        t = routineTimer.getTime()
        tThisFlip = win.getFutureFlipTime(clock=routineTimer)
        tThisFlipGlobal = win.getFutureFlipTime(clock=None)
        frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
        # update/draw components on each frame
        
        # *welcometext* updates
        
        # if welcometext is starting this frame...
        if welcometext.status == NOT_STARTED and tThisFlip >= 0.1-frameTolerance:
            # keep track of start time/frame for later
            welcometext.frameNStart = frameN  # exact frame index
            welcometext.tStart = t  # local t and not account for scr refresh
            welcometext.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(welcometext, 'tStartRefresh')  # time at next scr refresh
            # add timestamp to datafile
            thisExp.timestampOnFlip(win, 'welcometext.started')
            # update status
            welcometext.status = STARTED
            welcometext.setAutoDraw(True)
        
        # if welcometext is active this frame...
        if welcometext.status == STARTED:
            # update params
            pass
        
        # *forwardkey* updates
        waitOnFlip = False
        
        # if forwardkey is starting this frame...
        if forwardkey.status == NOT_STARTED and tThisFlip >= 3.5-frameTolerance:
            # keep track of start time/frame for later
            forwardkey.frameNStart = frameN  # exact frame index
            forwardkey.tStart = t  # local t and not account for scr refresh
            forwardkey.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(forwardkey, 'tStartRefresh')  # time at next scr refresh
            # add timestamp to datafile
            thisExp.timestampOnFlip(win, 'forwardkey.started')
            # update status
            forwardkey.status = STARTED
            # keyboard checking is just starting
            waitOnFlip = True
            win.callOnFlip(forwardkey.clock.reset)  # t=0 on next screen flip
            win.callOnFlip(forwardkey.clearEvents, eventType='keyboard')  # clear events on next screen flip
        if forwardkey.status == STARTED and not waitOnFlip:
            theseKeys = forwardkey.getKeys(keyList=['right'], ignoreKeys=["escape"], waitRelease=False)
            _forwardkey_allKeys.extend(theseKeys)
            if len(_forwardkey_allKeys):
                forwardkey.keys = _forwardkey_allKeys[-1].name  # just the last key pressed
                forwardkey.rt = _forwardkey_allKeys[-1].rt
                forwardkey.duration = _forwardkey_allKeys[-1].duration
                # a response ends the routine
                continueRoutine = False
        
        # *advance* updates
        
        # if advance is starting this frame...
        if advance.status == NOT_STARTED and tThisFlip >= 3.5-frameTolerance:
            # keep track of start time/frame for later
            advance.frameNStart = frameN  # exact frame index
            advance.tStart = t  # local t and not account for scr refresh
            advance.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(advance, 'tStartRefresh')  # time at next scr refresh
            # add timestamp to datafile
            thisExp.timestampOnFlip(win, 'advance.started')
            # update status
            advance.status = STARTED
            advance.setAutoDraw(True)
        
        # if advance is active this frame...
        if advance.status == STARTED:
            # update params
            pass
        
        # check for quit (typically the Esc key)
        if defaultKeyboard.getKeys(keyList=["escape"]):
            thisExp.status = FINISHED
        if thisExp.status == FINISHED or endExpNow:
            endExperiment(thisExp, inputs=inputs, win=win)
            return
        
        # check if all components have finished
        if not continueRoutine:  # a component has requested a forced-end of Routine
            routineForceEnded = True
            break
        continueRoutine = False  # will revert to True if at least one component still running
        for thisComponent in welcome_screenComponents:
            if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                continueRoutine = True
                break  # at least one component has not yet finished
        
        # refresh the screen
        if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
            win.flip()
    
    # --- Ending Routine "welcome_screen" ---
    for thisComponent in welcome_screenComponents:
        if hasattr(thisComponent, "setAutoDraw"):
            thisComponent.setAutoDraw(False)
    thisExp.addData('welcome_screen.stopped', globalClock.getTime())
    # check responses
    if forwardkey.keys in ['', [], None]:  # No response was made
        forwardkey.keys = None
    thisExp.addData('forwardkey.keys',forwardkey.keys)
    if forwardkey.keys != None:  # we had a response
        thisExp.addData('forwardkey.rt', forwardkey.rt)
        thisExp.addData('forwardkey.duration', forwardkey.duration)
    thisExp.nextEntry()
    # the Routine "welcome_screen" was not non-slip safe, so reset the non-slip timer
    routineTimer.reset()
    
    # --- Prepare to start Routine "warning_prep" ---
    continueRoutine = True
    # update component parameters for each repeat
    thisExp.addData('warning_prep.started', globalClock.getTime())
    forwardkey2.keys = []
    forwardkey2.rt = []
    _forwardkey2_allKeys = []
    # keep track of which components have finished
    warning_prepComponents = [warningtime, forwardkey2, advance2]
    for thisComponent in warning_prepComponents:
        thisComponent.tStart = None
        thisComponent.tStop = None
        thisComponent.tStartRefresh = None
        thisComponent.tStopRefresh = None
        if hasattr(thisComponent, 'status'):
            thisComponent.status = NOT_STARTED
    # reset timers
    t = 0
    _timeToFirstFrame = win.getFutureFlipTime(clock="now")
    frameN = -1
    
    # --- Run Routine "warning_prep" ---
    routineForceEnded = not continueRoutine
    while continueRoutine:
        # get current time
        t = routineTimer.getTime()
        tThisFlip = win.getFutureFlipTime(clock=routineTimer)
        tThisFlipGlobal = win.getFutureFlipTime(clock=None)
        frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
        # update/draw components on each frame
        
        # *warningtime* updates
        
        # if warningtime is starting this frame...
        if warningtime.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
            # keep track of start time/frame for later
            warningtime.frameNStart = frameN  # exact frame index
            warningtime.tStart = t  # local t and not account for scr refresh
            warningtime.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(warningtime, 'tStartRefresh')  # time at next scr refresh
            # add timestamp to datafile
            thisExp.timestampOnFlip(win, 'warningtime.started')
            # update status
            warningtime.status = STARTED
            warningtime.setAutoDraw(True)
        
        # if warningtime is active this frame...
        if warningtime.status == STARTED:
            # update params
            pass
        
        # *forwardkey2* updates
        waitOnFlip = False
        
        # if forwardkey2 is starting this frame...
        if forwardkey2.status == NOT_STARTED and tThisFlip >= 3.5-frameTolerance:
            # keep track of start time/frame for later
            forwardkey2.frameNStart = frameN  # exact frame index
            forwardkey2.tStart = t  # local t and not account for scr refresh
            forwardkey2.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(forwardkey2, 'tStartRefresh')  # time at next scr refresh
            # add timestamp to datafile
            thisExp.timestampOnFlip(win, 'forwardkey2.started')
            # update status
            forwardkey2.status = STARTED
            # keyboard checking is just starting
            waitOnFlip = True
            win.callOnFlip(forwardkey2.clock.reset)  # t=0 on next screen flip
            win.callOnFlip(forwardkey2.clearEvents, eventType='keyboard')  # clear events on next screen flip
        if forwardkey2.status == STARTED and not waitOnFlip:
            theseKeys = forwardkey2.getKeys(keyList=['right'], ignoreKeys=["escape"], waitRelease=False)
            _forwardkey2_allKeys.extend(theseKeys)
            if len(_forwardkey2_allKeys):
                forwardkey2.keys = _forwardkey2_allKeys[-1].name  # just the last key pressed
                forwardkey2.rt = _forwardkey2_allKeys[-1].rt
                forwardkey2.duration = _forwardkey2_allKeys[-1].duration
                # a response ends the routine
                continueRoutine = False
        
        # *advance2* updates
        
        # if advance2 is starting this frame...
        if advance2.status == NOT_STARTED and tThisFlip >= 3.5-frameTolerance:
            # keep track of start time/frame for later
            advance2.frameNStart = frameN  # exact frame index
            advance2.tStart = t  # local t and not account for scr refresh
            advance2.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(advance2, 'tStartRefresh')  # time at next scr refresh
            # add timestamp to datafile
            thisExp.timestampOnFlip(win, 'advance2.started')
            # update status
            advance2.status = STARTED
            advance2.setAutoDraw(True)
        
        # if advance2 is active this frame...
        if advance2.status == STARTED:
            # update params
            pass
        
        # check for quit (typically the Esc key)
        if defaultKeyboard.getKeys(keyList=["escape"]):
            thisExp.status = FINISHED
        if thisExp.status == FINISHED or endExpNow:
            endExperiment(thisExp, inputs=inputs, win=win)
            return
        
        # check if all components have finished
        if not continueRoutine:  # a component has requested a forced-end of Routine
            routineForceEnded = True
            break
        continueRoutine = False  # will revert to True if at least one component still running
        for thisComponent in warning_prepComponents:
            if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                continueRoutine = True
                break  # at least one component has not yet finished
        
        # refresh the screen
        if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
            win.flip()
    
    # --- Ending Routine "warning_prep" ---
    for thisComponent in warning_prepComponents:
        if hasattr(thisComponent, "setAutoDraw"):
            thisComponent.setAutoDraw(False)
    thisExp.addData('warning_prep.stopped', globalClock.getTime())
    # check responses
    if forwardkey2.keys in ['', [], None]:  # No response was made
        forwardkey2.keys = None
    thisExp.addData('forwardkey2.keys',forwardkey2.keys)
    if forwardkey2.keys != None:  # we had a response
        thisExp.addData('forwardkey2.rt', forwardkey2.rt)
        thisExp.addData('forwardkey2.duration', forwardkey2.duration)
    thisExp.nextEntry()
    # the Routine "warning_prep" was not non-slip safe, so reset the non-slip timer
    routineTimer.reset()
    
    # --- Prepare to start Routine "Experiment1_countdown" ---
    continueRoutine = True
    # update component parameters for each repeat
    thisExp.addData('Experiment1_countdown.started', globalClock.getTime())
    # Run 'Begin Routine' code from First_grid_code
    import random
    from openpyxl import load_workbook, Workbook
    import psychopy.event
    import psychopy.core
    import psychopy.data
    import psychopy.visual
    
    # Load image filenames from Excel file
    workbook = load_workbook('onion_stimuli.xlsx')
    worksheet = workbook.active
    image_filenames = [cell.value for cell in worksheet['A'][1:]]  # Get values from column A, starting from row 2
    random.shuffle(image_filenames)
    
    # Find the index of the correct image
    for i, filename in enumerate(image_filenames):
        if filename.endswith("onionFace.jpeg"):
            correct_index = i
            correct_answer = filename
            break
    else:
        raise ValueError("'onionFace.jpeg' not found in the image filenames.")
    
    # Create the 4x4 grid of images with tighter spacing and center it on the screen
    positions = [[-0.325 + col * 0.1625, 0.325 - row * 0.1625] for row in range(4) for col in range(4)]
    
    # Initialize data file
    thisExp = psychopy.data.ExperimentHandler(name="Experiment", savePickle=False, saveWideText=True, dataFileName="experiment_data")
    
    # Initialize a boolean to store the correctness of the response
    response_correct = False
    
    # Initialize mouse
    mouse = psychopy.event.Mouse()
    
    # Create the 4x4 grid of ImageStim objects
    OnionGrid = [psychopy.visual.ImageStim(win=win, image=filename, pos=positions[i], size=(0.15, 0.15)) for i, filename in enumerate(image_filenames)]
    
    # Add a timer
    timer = psychopy.core.CountdownTimer(10)  # 10 seconds timer
    
    # Main experiment loop
    while timer.getTime() > 0:
        # Draw the grid of images
        for img in OnionGrid:
            img.draw()
        win.flip()
    
        # Check for mouse click
        for i, image in enumerate(OnionGrid):
            if mouse.getPressed()[0] and image.contains(mouse):
                if image_filenames[i] == correct_answer:
                    response_correct = True
                else:
                    response_correct = False
                # Re-draw the images
                for img in OnionGrid:
                    img.draw()
                win.flip()
                psychopy.core.wait(1)
                break
        else:
            continue
        break
    
    # If no response, record it as incorrect
    if not response_correct:
        response_correct = False
        for img in OnionGrid:
            img.draw()
        win.flip()
        psychopy.core.wait(1)
    
    # Save the data for the Onion Grid experiment
    thisExp.addData('Onion Response Correctness', response_correct)
    thisExp.nextEntry()  # Move to the next row in the data file
    
    # End routine
    win.flip()  # Clear the screen
    # keep track of which components have finished
    Experiment1_countdownComponents = [OnionGrid]
    for image in OnionGrid:
        image.tStart = None
        image.tStop = None
        image.tStartRefresh = None
        image.tStopRefresh = None
        if hasattr(image, 'status'):
            image.status = NOT_STARTED
    # reset timers
    t = 0
    _timeToFirstFrame = win.getFutureFlipTime(clock="now")
    frameN = -1
    
    # --- Run Routine "Experiment1_countdown" ---
    routineForceEnded = not continueRoutine
    while continueRoutine:
        # get current time
        t = routineTimer.getTime()
        tThisFlip = win.getFutureFlipTime(clock=routineTimer)
        tThisFlipGlobal = win.getFutureFlipTime(clock=None)
        frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
        # update/draw components on each frame
        for image in OnionGrid:
            #if the image is starting this frame...
            if image.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                # keep track of start time/frame for later
                image.frameNStart = frameN  # exact frame index
                image.tStart = t  # local t and not account for scr refresh
                image.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(image, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'OnionGrid.started')
                # update status
                image.status = STARTED
                image.setAutoDraw(True)
        
            # if OnionGrid is active this frame...
            if image.status == STARTED:
                # update params
                pass
        
        # check if all components have finished
        if not continueRoutine:  # a component has requested a forced-end of Routine
            routineForceEnded = True
            break
        continueRoutine = False  # will revert to True if at least one component still running
        for image in Experiment1_countdownComponents:
            if hasattr(image, "status") and image.status != FINISHED:
                continueRoutine = True
                break  # at least one component has not yet finished
        
        # refresh the screen
        if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
            win.flip()
    
    # --- Ending Routine "Experiment1_countdown" ---
    for image in OnionGrid:
        if hasattr(image, "setAutoDraw"):
            image.setAutoDraw(False)
    thisExp.addData('Experiment1_countdown.stopped', globalClock.getTime())
    # the Routine "Experiment1_countdown" was not non-slip safe, so reset the non-slip timer
    routineTimer.reset()
    
    # --- Prepare to start Routine "second_prep" ---
    continueRoutine = True
    # update component parameters for each repeat
    thisExp.addData('second_prep.started', globalClock.getTime())
    forwardkey3.keys = []
    forwardkey3.rt = []
    _forwardkey3_allKeys = []
    # keep track of which components have finished
    second_prepComponents = [secondtestwarning, forwardkey3, advance3]
    for thisComponent in second_prepComponents:
        thisComponent.tStart = None
        thisComponent.tStop = None
        thisComponent.tStartRefresh = None
        thisComponent.tStopRefresh = None
        if hasattr(thisComponent, 'status'):
            thisComponent.status = NOT_STARTED
    # reset timers
    t = 0
    _timeToFirstFrame = win.getFutureFlipTime(clock="now")
    frameN = -1
    
    # --- Run Routine "second_prep" ---
    routineForceEnded = not continueRoutine
    while continueRoutine:
        # get current time
        t = routineTimer.getTime()
        tThisFlip = win.getFutureFlipTime(clock=routineTimer)
        tThisFlipGlobal = win.getFutureFlipTime(clock=None)
        frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
        # update/draw components on each frame
        
        # *secondtestwarning* updates
        
        # if secondtestwarning is starting this frame...
        if secondtestwarning.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
            # keep track of start time/frame for later
            secondtestwarning.frameNStart = frameN  # exact frame index
            secondtestwarning.tStart = t  # local t and not account for scr refresh
            secondtestwarning.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(secondtestwarning, 'tStartRefresh')  # time at next scr refresh
            # add timestamp to datafile
            thisExp.timestampOnFlip(win, 'secondtestwarning.started')
            # update status
            secondtestwarning.status = STARTED
            secondtestwarning.setAutoDraw(True)
        
        # if secondtestwarning is active this frame...
        if secondtestwarning.status == STARTED:
            # update params
            pass
        
        # *forwardkey3* updates
        waitOnFlip = False
        
        # if forwardkey3 is starting this frame...
        if forwardkey3.status == NOT_STARTED and tThisFlip >= 3.5-frameTolerance:
            # keep track of start time/frame for later
            forwardkey3.frameNStart = frameN  # exact frame index
            forwardkey3.tStart = t  # local t and not account for scr refresh
            forwardkey3.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(forwardkey3, 'tStartRefresh')  # time at next scr refresh
            # add timestamp to datafile
            thisExp.timestampOnFlip(win, 'forwardkey3.started')
            # update status
            forwardkey3.status = STARTED
            # keyboard checking is just starting
            waitOnFlip = True
            win.callOnFlip(forwardkey3.clock.reset)  # t=0 on next screen flip
            win.callOnFlip(forwardkey3.clearEvents, eventType='keyboard')  # clear events on next screen flip
        if forwardkey3.status == STARTED and not waitOnFlip:
            theseKeys = forwardkey3.getKeys(keyList=['right'], ignoreKeys=["escape"], waitRelease=False)
            _forwardkey3_allKeys.extend(theseKeys)
            if len(_forwardkey3_allKeys):
                forwardkey3.keys = _forwardkey3_allKeys[-1].name  # just the last key pressed
                forwardkey3.rt = _forwardkey3_allKeys[-1].rt
                forwardkey3.duration = _forwardkey3_allKeys[-1].duration
                # a response ends the routine
                continueRoutine = False
        
        # *advance3* updates
        
        # if advance3 is starting this frame...
        if advance3.status == NOT_STARTED and tThisFlip >= 3.5-frameTolerance:
            # keep track of start time/frame for later
            advance3.frameNStart = frameN  # exact frame index
            advance3.tStart = t  # local t and not account for scr refresh
            advance3.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(advance3, 'tStartRefresh')  # time at next scr refresh
            # add timestamp to datafile
            thisExp.timestampOnFlip(win, 'advance3.started')
            # update status
            advance3.status = STARTED
            advance3.setAutoDraw(True)
        
        # if advance3 is active this frame...
        if advance3.status == STARTED:
            # update params
            pass
        
        # check for quit (typically the Esc key)
        if defaultKeyboard.getKeys(keyList=["escape"]):
            thisExp.status = FINISHED
        if thisExp.status == FINISHED or endExpNow:
            endExperiment(thisExp, inputs=inputs, win=win)
            return
        
        # check if all components have finished
        if not continueRoutine:  # a component has requested a forced-end of Routine
            routineForceEnded = True
            break
        continueRoutine = False  # will revert to True if at least one component still running
        for thisComponent in second_prepComponents:
            if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                continueRoutine = True
                break  # at least one component has not yet finished
        
        # refresh the screen
        if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
            win.flip()
    
    # --- Ending Routine "second_prep" ---
    for thisComponent in second_prepComponents:
        if hasattr(thisComponent, "setAutoDraw"):
            thisComponent.setAutoDraw(False)
    thisExp.addData('second_prep.stopped', globalClock.getTime())
    # check responses
    if forwardkey3.keys in ['', [], None]:  # No response was made
        forwardkey3.keys = None
    thisExp.addData('forwardkey3.keys',forwardkey3.keys)
    if forwardkey3.keys != None:  # we had a response
        thisExp.addData('forwardkey3.rt', forwardkey3.rt)
        thisExp.addData('forwardkey3.duration', forwardkey3.duration)
    thisExp.nextEntry()
    # the Routine "second_prep" was not non-slip safe, so reset the non-slip timer
    routineTimer.reset()
    
    # --- Prepare to start Routine "Experiement2_countdown" ---
    continueRoutine = True
    # update component parameters for each repeat
    thisExp.addData('Experiement2_countdown.started', globalClock.getTime())
    # Run 'Begin Routine' code from memory_timer
    from psychopy import visual, core
    
    # Assuming you already have a window created and assigned to the variable `win`
    
    # Load the PNG image
    image_path = '24memorytest.png'
    image = visual.ImageStim(win=win, image=image_path, pos=(0, 0), size=(0.9, 0.7))
    
    # Create a text stimulus for the countdown timer
    timer_text = visual.TextStim(win=win, text='0:35', color='white', height=0.05, pos=(-0.5, 0.4))
    
    # Start the countdown timer
    timer = core.Clock()
    total_time = 35
    
    while timer.getTime() <= total_time:
        # Draw the image
        image.draw()
        
        # Draw the timer text
        remaining_time = total_time - timer.getTime()
        minutes = int(remaining_time // 60)
        seconds = int(remaining_time % 60)
        timer_text.text = '{:01d}:{:02d}'.format(minutes, seconds)
        timer_text.draw()
        
        # Update the window
        win.flip()
        
        core.wait(1)  # Update every second
    
    # keep track of which components have finished
    Experiement2_countdownComponents = []
    for thisComponent in Experiement2_countdownComponents:
        thisComponent.tStart = None
        thisComponent.tStop = None
        thisComponent.tStartRefresh = None
        thisComponent.tStopRefresh = None
        if hasattr(thisComponent, 'status'):
            thisComponent.status = NOT_STARTED
    # reset timers
    t = 0
    _timeToFirstFrame = win.getFutureFlipTime(clock="now")
    frameN = -1
    
    # --- Run Routine "Experiement2_countdown" ---
    routineForceEnded = not continueRoutine
    while continueRoutine and routineTimer.getTime() < 35.0:
        # get current time
        t = routineTimer.getTime()
        tThisFlip = win.getFutureFlipTime(clock=routineTimer)
        tThisFlipGlobal = win.getFutureFlipTime(clock=None)
        frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
        # update/draw components on each frame
        # is it time to end the Routine? (based on local clock)
        if tThisFlip > 35.0-frameTolerance:
            continueRoutine = False
        
        # check for quit (typically the Esc key)
        if defaultKeyboard.getKeys(keyList=["escape"]):
            thisExp.status = FINISHED
        if thisExp.status == FINISHED or endExpNow:
            endExperiment(thisExp, inputs=inputs, win=win)
            return
        
        # check if all components have finished
        if not continueRoutine:  # a component has requested a forced-end of Routine
            routineForceEnded = True
            break
        continueRoutine = False  # will revert to True if at least one component still running
        for thisComponent in Experiement2_countdownComponents:
            if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                continueRoutine = True
                break  # at least one component has not yet finished
        
        # refresh the screen
        if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
            win.flip()
    
    # --- Ending Routine "Experiement2_countdown" ---
    for thisComponent in Experiement2_countdownComponents:
        if hasattr(thisComponent, "setAutoDraw"):
            thisComponent.setAutoDraw(False)
    thisExp.addData('Experiement2_countdown.stopped', globalClock.getTime())
    # using non-slip timing so subtract the expected duration of this Routine (unless ended on request)
    if routineForceEnded:
        routineTimer.reset()
    else:
        routineTimer.addTime(-35.000000)
    
    # --- Prepare to start Routine "Experiement2_instruction" ---
    continueRoutine = True
    # update component parameters for each repeat
    thisExp.addData('Experiement2_instruction.started', globalClock.getTime())
    forwardkey4.keys = []
    forwardkey4.rt = []
    _forwardkey4_allKeys = []
    # keep track of which components have finished
    Experiement2_instructionComponents = [selectioninstruction, forwardkey4, advance4]
    for thisComponent in Experiement2_instructionComponents:
        thisComponent.tStart = None
        thisComponent.tStop = None
        thisComponent.tStartRefresh = None
        thisComponent.tStopRefresh = None
        if hasattr(thisComponent, 'status'):
            thisComponent.status = NOT_STARTED
    # reset timers
    t = 0
    _timeToFirstFrame = win.getFutureFlipTime(clock="now")
    frameN = -1
    
    # --- Run Routine "Experiement2_instruction" ---
    routineForceEnded = not continueRoutine
    while continueRoutine:
        # get current time
        t = routineTimer.getTime()
        tThisFlip = win.getFutureFlipTime(clock=routineTimer)
        tThisFlipGlobal = win.getFutureFlipTime(clock=None)
        frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
        # update/draw components on each frame
        
        # *selectioninstruction* updates
        
        # if selectioninstruction is starting this frame...
        if selectioninstruction.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
            # keep track of start time/frame for later
            selectioninstruction.frameNStart = frameN  # exact frame index
            selectioninstruction.tStart = t  # local t and not account for scr refresh
            selectioninstruction.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(selectioninstruction, 'tStartRefresh')  # time at next scr refresh
            # add timestamp to datafile
            thisExp.timestampOnFlip(win, 'selectioninstruction.started')
            # update status
            selectioninstruction.status = STARTED
            selectioninstruction.setAutoDraw(True)
        
        # if selectioninstruction is active this frame...
        if selectioninstruction.status == STARTED:
            # update params
            pass
        
        # *forwardkey4* updates
        waitOnFlip = False
        
        # if forwardkey4 is starting this frame...
        if forwardkey4.status == NOT_STARTED and tThisFlip >= 3.5-frameTolerance:
            # keep track of start time/frame for later
            forwardkey4.frameNStart = frameN  # exact frame index
            forwardkey4.tStart = t  # local t and not account for scr refresh
            forwardkey4.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(forwardkey4, 'tStartRefresh')  # time at next scr refresh
            # add timestamp to datafile
            thisExp.timestampOnFlip(win, 'forwardkey4.started')
            # update status
            forwardkey4.status = STARTED
            # keyboard checking is just starting
            waitOnFlip = True
            win.callOnFlip(forwardkey4.clock.reset)  # t=0 on next screen flip
            win.callOnFlip(forwardkey4.clearEvents, eventType='keyboard')  # clear events on next screen flip
        if forwardkey4.status == STARTED and not waitOnFlip:
            theseKeys = forwardkey4.getKeys(keyList=['right'], ignoreKeys=["escape"], waitRelease=False)
            _forwardkey4_allKeys.extend(theseKeys)
            if len(_forwardkey4_allKeys):
                forwardkey4.keys = _forwardkey4_allKeys[-1].name  # just the last key pressed
                forwardkey4.rt = _forwardkey4_allKeys[-1].rt
                forwardkey4.duration = _forwardkey4_allKeys[-1].duration
                # a response ends the routine
                continueRoutine = False
        
        # *advance4* updates
        
        # if advance4 is starting this frame...
        if advance4.status == NOT_STARTED and tThisFlip >= 3.5-frameTolerance:
            # keep track of start time/frame for later
            advance4.frameNStart = frameN  # exact frame index
            advance4.tStart = t  # local t and not account for scr refresh
            advance4.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(advance4, 'tStartRefresh')  # time at next scr refresh
            # add timestamp to datafile
            thisExp.timestampOnFlip(win, 'advance4.started')
            # update status
            advance4.status = STARTED
            advance4.setAutoDraw(True)
        
        # if advance4 is active this frame...
        if advance4.status == STARTED:
            # update params
            pass
        
        # check for quit (typically the Esc key)
        if defaultKeyboard.getKeys(keyList=["escape"]):
            thisExp.status = FINISHED
        if thisExp.status == FINISHED or endExpNow:
            endExperiment(thisExp, inputs=inputs, win=win)
            return
        
        # check if all components have finished
        if not continueRoutine:  # a component has requested a forced-end of Routine
            routineForceEnded = True
            break
        continueRoutine = False  # will revert to True if at least one component still running
        for thisComponent in Experiement2_instructionComponents:
            if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                continueRoutine = True
                break  # at least one component has not yet finished
        
        # refresh the screen
        if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
            win.flip()
    
    # --- Ending Routine "Experiement2_instruction" ---
    for thisComponent in Experiement2_instructionComponents:
        if hasattr(thisComponent, "setAutoDraw"):
            thisComponent.setAutoDraw(False)
    thisExp.addData('Experiement2_instruction.stopped', globalClock.getTime())
    # check responses
    if forwardkey4.keys in ['', [], None]:  # No response was made
        forwardkey4.keys = None
    thisExp.addData('forwardkey4.keys',forwardkey4.keys)
    if forwardkey4.keys != None:  # we had a response
        thisExp.addData('forwardkey4.rt', forwardkey4.rt)
        thisExp.addData('forwardkey4.duration', forwardkey4.duration)
    thisExp.nextEntry()
    # the Routine "Experiement2_instruction" was not non-slip safe, so reset the non-slip timer
    routineTimer.reset()
    
    # --- Prepare to start Routine "Experiment2_imageselection" ---
    continueRoutine = True
    # update component parameters for each repeat
    thisExp.addData('Experiment2_imageselection.started', globalClock.getTime())
    # Run 'Begin Routine' code from Second_grid_code
    import random
    from openpyxl import load_workbook, Workbook
    import psychopy.event
    import psychopy.core
    import psychopy.data
    import psychopy.visual
    
    # Load or create the Excel file
    try:
        workbook = load_workbook('experiment_data.xlsx')
        worksheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        worksheet = workbook.active
    
    # Load image filenames from Excel file
    workbook = load_workbook('tree_stimuli.xlsx')
    worksheet = workbook.active
    image_filenames = [cell.value for cell in worksheet['A'][1:]]  # Get values from column A, starting from row 2
    random.shuffle(image_filenames)
    
    # Prepare correct answer lists
    correct_answers_treeface = [
        "Timages/TreeFace1.jpeg",
        "Timages/TreeFace3.jpeg",
        "Timages/TreeFace4.jpeg",
        "Timages/TreeFace5.jpeg",
        "Timages/TreeFace9.jpeg",
        "Timages/TreeFace11.jpeg",
        "Timages/TreeFace13.jpeg",
        "Timages/TreeFace14.jpeg",
        "Timages/TreeFace15.jpeg",
        "Timages/TreeFace18.jpeg",
        "Timages/TreeFace23.jpeg",
        "Timages/TreeFace24.jpeg"
    ]
    
    correct_answers_tree = [
        "Timages/Tree1.jpeg",
        "Timages/Tree2.jpeg",
        "Timages/Tree3.jpeg",
        "Timages/Tree10.jpeg",
        "Timages/Tree11.jpeg",
        "Timages/Tree12.jpeg",
        "Timages/Tree14.jpeg",
        "Timages/Tree15.jpeg",
        "Timages/Tree16.jpeg",
        "Timages/Tree17.jpeg",
        "Timages/Tree21.jpeg",
        "Timages/Tree22.jpeg"
    ]
    
    # Create the 6x8 grid of images with tighter spacing and center it on the screen
    positions = [[-0.6 + col * 0.12, 0.3 - row * 0.12] for row in range(6) for col in range(8)]
    
    # Initialize image components with slightly smaller size and centered position
    TreeGrid = [[psychopy.visual.ImageStim(win=win, name='TreeGrid', image=None,
                        pos=positions[row * 8 + col], size=(0.1, 0.1)) for col in range(8)] for row in range(6)]
    
    # Set images in the grid
    for i, image_file in enumerate(image_filenames):
        row = i // 8
        col = i % 8
        TreeGrid[row][col].setImage(image_file)
    
    # Set up mouse event handling
    mouse = psychopy.event.Mouse(win=win)
    
    # Initialize the number of correct responses
    num_correct_treeface = 0
    num_correct_tree = 0
    
    # Main experiment loop
    participant_choices = []
    while continueRoutine and len(participant_choices) < 24:
        # Draw all images
        for row in range(6):
            for col in range(8):
                TreeGrid[row][col].draw()
        # Draw mouse cursor
        pos = mouse.getPos()
        buttons = mouse.getPressed()
        mouse_cursor = psychopy.visual.Circle(win, radius=0.001, fillColor='red', lineColor='red')
        mouse_cursor.pos = pos
        mouse_cursor.draw()
        win.flip()
        # Check for mouse click
        for row in range(6):
            for col in range(8):
                if TreeGrid[row][col].contains(pos) and buttons[0]:
                    participant_choice = TreeGrid[row][col].image
                    participant_choices.append(participant_choice)
                    if participant_choice in correct_answers_treeface:
                        num_correct_treeface += 1
                    elif participant_choice in correct_answers_tree:
                        num_correct_tree += 1
                    TreeGrid[row][col].opacity = 0.3
                    break
    
    # Save the data for the Tree Grid experiment
    thisExp.addData('Tree Grid Correct Responses (TreeFace)', num_correct_treeface)
    thisExp.addData('Tree Grid Correct Responses (Tree)', num_correct_tree)
    
    # End routine
    psychopy.core.wait(2.5)  # Wait before clearing the screen
    win.flip()  # Clear the screen
    # keep track of which components have finished
    Experiment2_imageselectionComponents = TreeGrid
    for row in Experiment2_imageselectionComponents:
        for imaget in row:
            imaget.tStart = None
            imaget.tStop = None
            imaget.tStartRefresh = None
            imaget.tStopRefresh = None
            if hasattr(imaget, 'status'):
                imaget.status = NOT_STARTED
    #Set continueRoutine to False to move on to next routine
    continueRoutine = False
    routineTimer.reset()
    
    # --- Run Routine "Experiment2_imageselection" ---
    while continueRoutine:
        # get current time
        t = routineTimer.getTime()
        tThisFlip = win.getFutureFlipTime(clock=routineTimer)
        tThisFlipGlobal = win.getFutureFlipTime(clock=None)
        frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
        # update/draw components on each frame
        for row in TreeGrid:
            for imaget in row:
                # if TreeGrid is starting this frame...
                if imaget.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                    # keep track of start time/frame for later
                    imaget.frameNStart = frameN  # exact frame index
                    imaget.tStart = t  # local t and not account for scr refresh
                    imaget.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(imaget, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'TreeGrid.started')
                    # update status
                    imaget.status = STARTED
                    imaget.setAutoDraw(True)
                
                # if TreeGrid is active this frame...
                if imaget.status == STARTED:
                    # update params
                    pass
                
        # check if all components have finished
        if not continueRoutine:  # a component has requested a forced-end of Routine
            routineForceEnded = True
            break
        continueRoutine = False  # will revert to True if at least one component still running
        for row in TreeGrid:
            for imaget in row:
                if hasattr(imaget, "status") and imaget.status != FINISHED:
                    continueRoutine = True
                    break  # at least one component has not yet finished
        
        # refresh the screen
        if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
            win.flip()
    
    # --- Ending Routine "Experiment2_imageselection" ---
    for row in TreeGrid:
        for imaget in row:
            if hasattr(imaget, "setAutoDraw"):
                imaget.setAutoDraw(False)
    thisExp.addData('Experiment2_imageselection.stopped', globalClock.getTime())
    # the Routine "Experiment2_imageselection" was not non-slip safe, so reset the non-slip timer
    routineTimer.reset()
    
    # --- Prepare to start Routine "End" ---
    continueRoutine = True
    # update component parameters for each repeat
    thisExp.addData('End.started', globalClock.getTime())
    endkey.keys = []
    endkey.rt = []
    _endkey_allKeys = []
    # keep track of which components have finished
    EndComponents = [endtext, endkey, endadvance]
    for thisComponent in EndComponents:
        thisComponent.tStart = None
        thisComponent.tStop = None
        thisComponent.tStartRefresh = None
        thisComponent.tStopRefresh = None
        if hasattr(thisComponent, 'status'):
            thisComponent.status = NOT_STARTED
    # reset timers
    t = 0
    _timeToFirstFrame = win.getFutureFlipTime(clock="now")
    frameN = -1
    
    # --- Run Routine "End" ---
    routineForceEnded = not continueRoutine
    while continueRoutine:
        # get current time
        t = routineTimer.getTime()
        tThisFlip = win.getFutureFlipTime(clock=routineTimer)
        tThisFlipGlobal = win.getFutureFlipTime(clock=None)
        frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
        # update/draw components on each frame
        
        # *endtext* updates
        
        # if endtext is starting this frame...
        if endtext.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
            # keep track of start time/frame for later
            endtext.frameNStart = frameN  # exact frame index
            endtext.tStart = t  # local t and not account for scr refresh
            endtext.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(endtext, 'tStartRefresh')  # time at next scr refresh
            # add timestamp to datafile
            thisExp.timestampOnFlip(win, 'endtext.started')
            # update status
            endtext.status = STARTED
            endtext.setAutoDraw(True)
        
        # if endtext is active this frame...
        if endtext.status == STARTED:
            # update params
            pass
        
        # *endkey* updates
        waitOnFlip = False
        
        # if endkey is starting this frame...
        if endkey.status == NOT_STARTED and tThisFlip >= 3.5-frameTolerance:
            # keep track of start time/frame for later
            endkey.frameNStart = frameN  # exact frame index
            endkey.tStart = t  # local t and not account for scr refresh
            endkey.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(endkey, 'tStartRefresh')  # time at next scr refresh
            # add timestamp to datafile
            thisExp.timestampOnFlip(win, 'endkey.started')
            # update status
            endkey.status = STARTED
            # keyboard checking is just starting
            waitOnFlip = True
            win.callOnFlip(endkey.clock.reset)  # t=0 on next screen flip
            win.callOnFlip(endkey.clearEvents, eventType='keyboard')  # clear events on next screen flip
        if endkey.status == STARTED and not waitOnFlip:
            theseKeys = endkey.getKeys(keyList=['right'], ignoreKeys=["escape"], waitRelease=False)
            _endkey_allKeys.extend(theseKeys)
            if len(_endkey_allKeys):
                endkey.keys = _endkey_allKeys[-1].name  # just the last key pressed
                endkey.rt = _endkey_allKeys[-1].rt
                endkey.duration = _endkey_allKeys[-1].duration
                # a response ends the routine
                continueRoutine = False
        
        # *endadvance* updates
        
        # if endadvance is starting this frame...
        if endadvance.status == NOT_STARTED and tThisFlip >= 3.5-frameTolerance:
            # keep track of start time/frame for later
            endadvance.frameNStart = frameN  # exact frame index
            endadvance.tStart = t  # local t and not account for scr refresh
            endadvance.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(endadvance, 'tStartRefresh')  # time at next scr refresh
            # add timestamp to datafile
            thisExp.timestampOnFlip(win, 'endadvance.started')
            # update status
            endadvance.status = STARTED
            endadvance.setAutoDraw(True)
        
        # if endadvance is active this frame...
        if endadvance.status == STARTED:
            # update params
            pass
        
        # check for quit (typically the Esc key)
        if defaultKeyboard.getKeys(keyList=["escape"]):
            thisExp.status = FINISHED
        if thisExp.status == FINISHED or endExpNow:
            endExperiment(thisExp, inputs=inputs, win=win)
            return
        
        # check if all components have finished
        if not continueRoutine:  # a component has requested a forced-end of Routine
            routineForceEnded = True
            break
        continueRoutine = False  # will revert to True if at least one component still running
        for thisComponent in EndComponents:
            if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                continueRoutine = True
                break  # at least one component has not yet finished
        
        # refresh the screen
        if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
            win.flip()
    
    # --- Ending Routine "End" ---
    for thisComponent in EndComponents:
        if hasattr(thisComponent, "setAutoDraw"):
            thisComponent.setAutoDraw(False)
    thisExp.addData('End.stopped', globalClock.getTime())
    # check responses
    if endkey.keys in ['', [], None]:  # No response was made
        endkey.keys = None
    thisExp.addData('endkey.keys',endkey.keys)
    if endkey.keys != None:  # we had a response
        thisExp.addData('endkey.rt', endkey.rt)
        thisExp.addData('endkey.duration', endkey.duration)
    thisExp.nextEntry()
    # the Routine "End" was not non-slip safe, so reset the non-slip timer
    routineTimer.reset()
    
    # mark experiment as finished
    endExperiment(thisExp, win=win, inputs=inputs)


def saveData(thisExp):
    """
    Save data from this experiment
    
    Parameters
    ==========
    thisExp : psychopy.data.ExperimentHandler
        Handler object for this experiment, contains the data to save and information about 
        where to save it to.
    """
    filename = thisExp.dataFileName
    # these shouldn't be strictly necessary (should auto-save)
    thisExp.saveAsWideText(filename + '.csv', delim='auto')
    thisExp.saveAsPickle(filename)


def endExperiment(thisExp, inputs=None, win=None):
    """
    End this experiment, performing final shut down operations.
    
    This function does NOT close the window or end the Python process - use `quit` for this.
    
    Parameters
    ==========
    thisExp : psychopy.data.ExperimentHandler
        Handler object for this experiment, contains the data to save and information about 
        where to save it to.
    inputs : dict
        Dictionary of input devices by name.
    win : psychopy.visual.Window
        Window for this experiment.
    """
    if win is not None:
        # remove autodraw from all current components
        win.clearAutoDraw()
        # Flip one final time so any remaining win.callOnFlip() 
        # and win.timeOnFlip() tasks get executed
        win.flip()
    # mark experiment handler as finished
    thisExp.status = FINISHED
    # shut down eyetracker, if there is one
    if inputs is not None:
        if 'eyetracker' in inputs and inputs['eyetracker'] is not None:
            inputs['eyetracker'].setConnectionState(False)


def quit(thisExp, win=None, inputs=None, thisSession=None):
    """
    Fully quit, closing the window and ending the Python process.
    
    Parameters
    ==========
    win : psychopy.visual.Window
        Window to close.
    inputs : dict
        Dictionary of input devices by name.
    thisSession : psychopy.session.Session or None
        Handle of the Session object this experiment is being run from, if any.
    """
    thisExp.abort()  # or data files will save again on exit
    # make sure everything is closed down
    if win is not None:
        # Flip one final time so any remaining win.callOnFlip() 
        # and win.timeOnFlip() tasks get executed before quitting
        win.flip()
        win.close()
    if inputs is not None:
        if 'eyetracker' in inputs and inputs['eyetracker'] is not None:
            inputs['eyetracker'].setConnectionState(False)
    if thisSession is not None:
        thisSession.stop()
    # terminate Python process
    core.quit()


# if running this experiment as a script...
if __name__ == '__main__':
    # call all functions in order
    expInfo = showExpInfoDlg(expInfo=expInfo)
    thisExp = setupData(expInfo=expInfo)
    logFile = setupLogging(filename=thisExp.dataFileName)
    win = setupWindow(expInfo=expInfo)
    inputs = setupInputs(expInfo=expInfo, thisExp=thisExp, win=win)
    run(
        expInfo=expInfo, 
        thisExp=thisExp, 
        win=win, 
        inputs=inputs
    )
    saveData(thisExp=thisExp)
    quit(thisExp=thisExp, win=win, inputs=inputs)
